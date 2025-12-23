import openpyxl as w
from openpyxl.styles import Font, Color
import pandas as pd
import Alim_bbdd_sah
ft = Font(color="FF0000")

def cargar_precios_web():
    return(Alim_bbdd_sah.table_to_df_e('SELECT * FROM v_brp_carreras_product'))

def crear_df():
    df = pd.DataFrame()
    df['referencia'] = None
    df['precio'] = None

def libro(ruta,hoja):
    df_precios = cargar_precios_web()
    wb = w.load_workbook(filename=ruta,data_only = True)
    sheet = wb[hoja]
    posicion = 2
    for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
        posicion = posicion +1
        for cell in row:
            try:
                series_precios = df_precios[df_precios['reference'] == str(cell.value)]['price'].reset_index()
                sheet['D'+str(posicion)]=series_precios['price'][0]
            except:
                cell.font = ft



    wb.save(filename='./salida_excel/carreras.xlsx')

    print(sheet)



#if __name__=="__main__":
